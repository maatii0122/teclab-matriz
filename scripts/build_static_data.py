from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Matriz general IPP completa actualizada.xlsx"
CLEAN_XLSX = ROOT / "Matriz limpia.xlsx"
PUBLIC_DATA = ROOT / "public" / "data" / "matriz.json"
PUBLIC_XLSX = ROOT / "public" / "Matriz limpia.xlsx"

FIELDS = [
    "PERIODO",
    "CAMPO DE FORMACIÓN",
    "CARGA HORARIA",
    "OBJETIVOS DE LA MATERIA",
    "CONTENIDOS MÍNIMOS",
    "BIBLIOGRAFÍA DE CONSULTA",
    "PRODUCCIÓN DE CONTENIDOS",
    "CARRERAS",
    "AÑO",
]

def clean_value(value) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def split_values(value) -> list[str]:
    text = clean_value(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def normalize_period(value) -> str | None:
    text = clean_value(value)
    if not text:
        return None
    compact = re.sub(r"\s+", "", text.upper())
    match = re.fullmatch(r"(\d)A/B", compact)
    if match:
        number = match.group(1)
        return f"{number}A / {number}B"
    match = re.fullmatch(r"(\d)([AB])", compact)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    return text


def period_values(value) -> list[str]:
    text = normalize_period(value)
    if not text:
        return []
    match = re.fullmatch(r"(\d)A\s*/\s*(\d)?B", text.upper())
    if match:
        number = match.group(1)
        return [f"{number}A", f"{number}B"]
    return split_values(text)


def unique_values(values, split: bool = False) -> list[str]:
    seen = set()
    result = []
    for value in values:
        parts = split_values(value) if split else [clean_value(value)]
        for part in parts:
            if part and part not in seen:
                seen.add(part)
                result.append(part)
    return result


def sort_text(value) -> str:
    text = clean_value(value) or ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char)).upper()


def build_clean_matrix(data: pd.DataFrame) -> pd.DataFrame:
    return data.sort_values("MATERIA", key=lambda values: values.map(sort_text), kind="mergesort")


def normalize_data(data: pd.DataFrame) -> pd.DataFrame:
    normalized = data.copy()
    normalized["PERIODO"] = normalized["PERIODO"].apply(normalize_period)
    return normalized


def records(data: pd.DataFrame) -> list[dict[str, str]]:
    normalized = data.fillna("")
    return normalized.to_dict(orient="records")


def metrics(data: pd.DataFrame) -> list[dict[str, str | int]]:
    total = len(data)
    result = []
    for field in FIELDS:
        filled = int(data[field].notna().sum())
        missing = total - filled
        percent = f"{((filled / total) * 100 if total else 0):.1f}%"
        result.append({"field": field, "filled": filled, "missing": missing, "percent": percent})
    return result


def write_formatted_excel(data: pd.DataFrame, path: Path) -> None:
    wrap_columns = {
        "OBJETIVOS DE LA MATERIA",
        "CONTENIDOS MÍNIMOS",
        "BIBLIOGRAFÍA DE CONSULTA",
        "PRODUCCIÓN DE CONTENIDOS",
        "ENUNCIADOS",
        "PRODUCCIÓN DE CONTENIDOS DE LA MATERIA",
    }

    data.to_excel(path, index=False, sheet_name="Matriz limpia")
    wb = load_workbook(path)
    ws = wb["Matriz limpia"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="03323D")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2E8")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    for column_idx, column_name in enumerate(data.columns, start=1):
        letter = get_column_letter(column_idx)
        sample_values = [str(column_name)] + [str(value) for value in data[column_name].dropna().head(100)]
        max_length = max(len(value) for value in sample_values) if sample_values else len(str(column_name))
        width = min(max(max_length + 2, 12), 55)
        if column_name in wrap_columns:
            width = min(max(width, 30), 55)
        ws.column_dimensions[letter].width = width

    wrap_letters = {
        get_column_letter(idx)
        for idx, name in enumerate(data.columns, start=1)
        if name in wrap_columns
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in wrap_letters)

    subject_column = list(data.columns).index("MATERIA") + 1
    merge_repeated_subjects(ws, subject_column)

    wb.save(path)


def merge_repeated_subjects(ws, column_idx: int) -> None:
    start_row = 2
    current_value = ws.cell(row=start_row, column=column_idx).value

    for row_idx in range(start_row + 1, ws.max_row + 2):
        next_value = ws.cell(row=row_idx, column=column_idx).value if row_idx <= ws.max_row else None
        if next_value == current_value:
            continue

        end_row = row_idx - 1
        if current_value and end_row > start_row:
            ws.merge_cells(
                start_row=start_row,
                start_column=column_idx,
                end_row=end_row,
                end_column=column_idx,
            )
            ws.cell(row=start_row, column=column_idx).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        start_row = row_idx
        current_value = next_value


def main() -> None:
    raw = normalize_data(pd.read_excel(SOURCE, engine="openpyxl"))
    clean = build_clean_matrix(raw)

    PUBLIC_DATA.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "columns": list(raw.columns),
        "fields": FIELDS,
        "rawRows": records(raw),
        "cleanRows": records(clean),
        "metrics": metrics(clean),
        "stats": {
            "rawRows": len(raw),
            "cleanRows": len(clean),
            "uniqueSubjects": int(raw["MATERIA"].nunique(dropna=True)),
            "careers": len(unique_values(raw["CARRERAS"], split=True)),
            "periods": len(unique_values(raw["PERIODO"].apply(period_values).explode(), split=False)),
        },
    }

    PUBLIC_DATA.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    write_formatted_excel(clean, CLEAN_XLSX)
    write_formatted_excel(clean, PUBLIC_XLSX)

    print(f"Wrote {PUBLIC_DATA.relative_to(ROOT)}")
    print(f"Wrote {CLEAN_XLSX.name}")
    print(f"Wrote {PUBLIC_XLSX.relative_to(ROOT)}")
    print(f"Rows: {len(raw)} raw -> {len(clean)} clean sorted rows")
    print(f"Unique subjects: {raw['MATERIA'].nunique(dropna=True)}")


if __name__ == "__main__":
    main()
